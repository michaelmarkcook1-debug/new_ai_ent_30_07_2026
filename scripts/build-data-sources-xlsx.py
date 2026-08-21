"""Build the data-sources register as a workbook for the pod.

WHY A SCRIPT AND NOT A ONE-OFF. docs/DATA-SOURCES.md is the source of truth and
this is a second surface onto it, so it has to be regenerable rather than
hand-assembled: a spreadsheet edited by hand drifts from the register the first
time either changes, and then the pod has two answers.

Every status column carries the date it was probed. A register that says a
source is healthy without saying when it last answered is the failure this
whole document exists to prevent, and it is exactly what happened here: the
markdown said the catalogue was live while its host had stopped resolving.

Usage:  python3 scripts/build-data-sources-xlsx.py [out.xlsx]
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "AI-Enterprise-data-sources.xlsx"

PROBED = "21 Aug 2026"
COMMIT = "8aa3657"

FONT = "Arial"
HEAD_FILL = PatternFill("solid", start_color="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SUB_FONT = Font(name=FONT, size=9, color="595959")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
MONO = Font(name="Consolas", size=9)

OK = PatternFill("solid", start_color="E2EFDA")
BAD = PatternFill("solid", start_color="FCE4E4")
WARN = PatternFill("solid", start_color="FFF2CC")
BAND = PatternFill("solid", start_color="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def header(ws, row, cols):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BOX
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def title(ws, text, sub):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws["A2"] = sub
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[1].height = 20


def write(ws, row, values, fills=None, wrap_cols=()):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = BODY
        cell.border = BOX
        cell.alignment = WRAP if i in wrap_cols else TOP
        if fills and i in fills:
            cell.fill = fills[i]


wb = Workbook()

# ─── 1. Every source ────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Sources"
title(
    ws,
    "AI Enterprise: data sources",
    f"Every upstream, its tier, and whether it answered. Probed {PROBED}. "
    f"Source of truth is docs/DATA-SOURCES.md at commit {COMMIT}; this sheet is a second surface onto it.",
)
header(
    ws,
    4,
    [
        "Source",
        "Tier",
        "What it gives us",
        "Endpoint or file",
        "Auth",
        "Status " + PROBED,
        "On failure",
        "Refresh",
    ],
)

ROWS = [
    # source, tier, gives, endpoint, auth, status, ok/bad/warn, on failure, refresh
    ("SEC EDGAR full-text", "Third-party live", "Company filings naming AI vendors. Industry breakdown by SIC code, from the regulator's own aggregation",
     "efts.sec.gov/LATEST/search-index", "None. User-Agent required by fair-access policy", "200, 0.5s, 36 KB", "ok",
     "Committed snapshot at data/adoption/disclosure-10-K.json", "Live per request, 5 min cache"),
    ("Federal Register", "Third-party live", "US federal AI rulemaking, roughly 1,500 documents",
     "federalregister.gov/api/v1/documents.json", "None", "200, 0.6s, 47 KB", "ok",
     "Panel renders nothing rather than stale", "Live per request"),
    ("Ranking engine (AIE v1)", "Proxied", "43 vendors, the 0 to 5 weighted assessment, market share, capabilities, reputation",
     "ranking-engine-red.vercel.app/api (10 whitelisted paths)", "None", "200, 0.4s, 74 KB", "ok",
     "12 recorded fixtures in fixtures/aie-live/", "5 min cache, 12 s timeout, 1 retry"),
    ("Anthropic API", "Third-party live", "The analyst voice and the Interrogate findings",
     "api.anthropic.com", "ANTHROPIC_API_KEY", "Live", "ok",
     "Scripted mode, zero LLM spend, stated on screen", "Per request, 24 h cache"),
    ("Google favicon", "Third-party live", "Vendor logos",
     "google.com/s2/favicons", "None", "301 redirect, works", "warn",
     "Blank SVG", "24 h cache"),
    ("BoardRadar (AnalystGenius)", "Proxied", "Live financials, reputation, governance and talent for listed companies",
     "ag-api-prod-calm-seastar-79.fly.dev/api/v1 (23 whitelisted prefixes)", "ANALYSTGENIUS_API_KEY, server-side only", "503 on every path. Host alive (/health 200) but /api/v1 returns 404", "bad",
     "88 recorded fixtures in fixtures/br/. Lane flips to error, no stale figure shown", "5 min cache, 60 req/min per IP"),
    ("Movement catalogue (Supabase)", "First-party live", "1,340 observations across model, vendor and market series. The only source we own end to end",
     "lmptnwqthldbficddtfn.supabase.co", "Publishable key to read, service key to write", "DEAD. Host does not resolve (NXDOMAIN). Every /api/catalogue/* returns 502", "bad",
     "No fallback. Panels render their own empty state", "Was on demand"),
    ("Clearbit logo", "Third-party live", "Vendor logos",
     "logo.clearbit.com", "None", "DEAD. DNS does not resolve. Retired after the HubSpot acquisition", "bad",
     "Fails in ~19 ms, cached 24 h, 1x1 blank SVG. No page breaks", "n/a"),
    ("Role library", "Bundled", "297 role profiles across 37 industries, 18 capabilities each",
     "lib/model-fit/data/roles.json (685 KB)", "n/a", "Snapshot, 2 Aug 2026", "warn",
     "n/a: it is a file", "Manual. Does not refresh"),
    ("Model catalogue", "Bundled", "330 priced and benchmarked models",
     "lib/model-fit/data/models.json (131 KB)", "n/a", "Snapshot, 2 Aug 2026", "warn",
     "n/a", "Manual"),
    ("Category assessment", "Bundled", "The 0 to 5 weighted composite per market, with every domain and its evidence grade",
     "fixtures/aie-live/category-rankings.json (373 KB)", "n/a", "Captured 17 Aug 2026", "warn",
     "n/a", "scripts/sync-category-rankings.mjs. Fails loudly rather than writing an empty ranking"),
    ("Privacy and IP Shield", "Cited", "14 model providers graded on training, retention, indemnity and residency, quoted from their own terms",
     "lib/shield/data.ts", "n/a", "Read by a human 14 Jul 2026. Re-read due at 30 days", "warn",
     "n/a", "Manual re-read. The page counts the days itself"),
    ("Workflow catalogue", "Bundled", "75 workflows across 15 sectors with risk tier, reliability bar and regulatory flags",
     "lib/aie/use-cases.ts", "n/a", "Curated", "warn", "n/a", "Manual"),
    ("Regulatory register", "Bundled", "Dated AI obligations by jurisdiction, each binding the provider, the deployer or both",
     "lib/aie/regulation/obligations.ts", "n/a", "EU dates re-verified 17 Aug 2026 against source", "ok",
     "n/a", "Manual, re-verified on change"),
]

FILLS = {"ok": OK, "bad": BAD, "warn": WARN}
r = 5
for src, tier, gives, ep, auth, status, flag, fail, refresh in ROWS:
    write(ws, r, [src, tier, gives, ep, auth, status, fail, refresh],
          fills={6: FILLS[flag]}, wrap_cols=(3, 4, 6, 7, 8))
    ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=4).font = MONO
    ws.row_dimensions[r].height = 46
    r += 1

widths(ws, {"A": 26, "B": 17, "C": 42, "D": 40, "E": 26, "F": 38, "G": 40, "H": 30})
ws.auto_filter.ref = f"A4:H{r - 1}"

# ─── 2. The tier rules ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Tiers and rules")
title(ws2, "How a figure is sourced, and how it is badged",
      "The distinction the whole product rests on. A bundled figure shown as live is the failure this prevents.")
header(ws2, 4, ["Tier", "Meaning", "Can it move?", "Badge on screen"])
TIERS = [
    ("First-party live", "We own the endpoint and the data behind it", "Yes, we control refresh", "LIVE"),
    ("Third-party live", "Someone else's API, fetched at request time", "Yes, on their cadence", "LIVE"),
    ("Proxied", "Someone else's API behind our whitelist and cache", "Only as fast as they refresh", "AIE LIVE"),
    ("Bundled", "A dated file in the repo", "No. A snapshot, and labelled as one", "AIE DATASET"),
    ("Cited", "A sentence quoted from a vendor's published terms, with the URL and the date a human read it", "No. Legal terms have no feed to poll", "CITED"),
    ("Derived", "Computed here from named inputs the reader can re-check", "Only when its inputs do", "DERIVED"),
]
r = 5
for t in TIERS:
    write(ws2, r, list(t), wrap_cols=(2, 3))
    ws2.cell(row=r, column=1).font = BOLD
    ws2.cell(row=r, column=4).font = MONO
    ws2.row_dimensions[r].height = 32
    r += 1

r += 1
ws2.cell(row=r, column=1, value="Two rules that govern every badge").font = Font(name=FONT, bold=True, size=11, color="1F3864")
r += 1
for rule in [
    "WORST LANE WINS. A panel combining live and sample data is badged sample. The badge describes the weakest input, never the strongest.",
    "A LANE IS NEVER ASSERTED BEFORE THE DATA ARRIVES. A LIVE badge over a spinner claims a fact about a response that has not happened. Components seed their source as null and render no badge until it resolves.",
    "Proxied responses carry the lane in the x-eai-source header. The badge is driven by that header and nothing else, so a wrong badge is almost never a component bug.",
]:
    c = ws2.cell(row=r, column=1, value=rule)
    c.font = BODY
    c.alignment = WRAP
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws2.row_dimensions[r].height = 30
    r += 1

widths(ws2, {"A": 22, "B": 62, "C": 34, "D": 18})

# ─── 3. What is broken ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("Open issues")
title(ws3, "Known issues", f"Probed {PROBED}. Two of these are new and are not yet in docs/DATA-SOURCES.md.")
header(ws3, 4, ["#", "Issue", "Severity", "Impact", "Status", "What it would take"])
ISSUES = [
    (1, "Movement catalogue host does not resolve (NXDOMAIN)", "HIGH", "A whole first-party tier gone. 1,340 observations unreachable. Every /api/catalogue/* 502s. AI Adoption's movement panel renders empty", "NEW, unresolved", "Confirm whether the Supabase project was deleted or paused. If deliberate, remove the series from the product and say so. If not, restore and re-run ingest:catalogue", "bad"),
    (2, "BoardRadar returns 503 on every path", "HIGH", "Six pages fall back to fixtures: Company View, Competitive Intel, Ecosystem Navigator, News Feed, Reputation Tracker, Security Desk", "NEW, unresolved", "Host is alive but /api/v1 404s. Check whether the API path moved or the key is not reaching production", "bad"),
    (3, "uptake is a static May 2026 seed", "HIGH (misleading)", "Its ordering puts OpenAI ahead of Anthropic, contradicted by Menlo and Ramp", "Mitigated", "Disclosed on-page, and the measured figures are shown above it. Nothing fresher exists behind it", "warn"),
    (4, "logo.clearbit.com retired", "LOW", "A missing logo, never an error", "Open", "Replace or remove at leisure", "warn"),
    (5, "/api/news returns 3.28 MB and ignores ?limit", "MEDIUM (bandwidth)", "Slow first load", "Mitigated", "24 h module TTL and a trim to 300 items", "warn"),
    (6, "OpenAI news headline still names Microsoft", "MEDIUM (accuracy)", "The News tab shows an upstream headline we know is wrong. Microsoft was not in that round", "Open", "Needs a correction printed beside it at render. Rewriting a recorded upstream response would falsify what the source said", "warn"),
    (7, "docs/DATA-SOURCES.md last verified 5 Aug", "MEDIUM", "It lists both dead sources above as healthy", "NEW", "Re-verify and restate, which this sheet is the evidence for", "warn"),
]
r = 5
for n, issue, sev, impact, status, todo, flag in ISSUES:
    write(ws3, r, [n, issue, sev, impact, status, todo], fills={3: FILLS[flag]}, wrap_cols=(2, 4, 5, 6))
    ws3.cell(row=r, column=2).font = BOLD
    ws3.row_dimensions[r].height = 52
    r += 1
widths(ws3, {"A": 5, "B": 38, "C": 18, "D": 46, "E": 18, "F": 50})
ws3.auto_filter.ref = f"A4:F{r - 1}"

# ─── 4. Page to source ──────────────────────────────────────────────────────
ws4 = wb.create_sheet("Page to source")
title(ws4, "What each page runs on", "SAMPLE counts read from rendered HTML, not from props. Four pages carry sample data; the rest carry none.")
header(ws4, 4, ["Page", "Question it answers", "Data behind it", "Lanes shown", "SAMPLE badges"])
PAGES = [
    ("/start", "Which question is yours", "Static card copy", "AIE LIVE, LIVE", 0),
    ("/pulse", "What changed today", "market-share, reputation, capabilities; judgement derived", "AIE, DERIVED", 0),
    ("/market-watch", "Who leads each category", "market-share.json", "AIE LIVE, AIE", 0),
    ("/ai-adoption", "Who is actually paying", "Menlo and Ramp curated, SEC disclosure, catalogue movement", "LIVE, AIE LIVE, AIE", 0),
    ("/peer-insights", "What firms like mine buy", "AIE uptake, a May 2026 seed", "AIE LIVE", 0),
    ("/financial-snapshot", "What vendors disclose about AI revenue", "SEC filings, disclosure ladder", "LIVE, AIE", 0),
    ("/competitive-intel", "Capability comparison", "capabilities.json, 470 vendor-capability rows, all 13 markets", "matrix lane, LIVE", 0),
    ("/vendor-view", "One vendor, read properly", "vendors, capabilities, reputation, composite", "AIE", 0),
    ("/reputation-tracker", "How buyers rate vendors", "reputation.json plus third-party block", "AIE, LIVE", 1),
    ("/alliances (Integrators)", "Which firms deliver which vendors", "alliances seed, 51 channel links", "AIE", 0),
    ("/market-view (ModelEngine)", "Which model for a role", "roles.json (297), models.json (330)", "DERIVED, AIE", 0),
    ("/price-performance", "What capability costs", "models.json", "AIE LIVE, AIE", 0),
    ("/workflow-shortlist", "Who to buy for a workflow", "75 workflows plus vendor index", "AIE LIVE, AIE", 0),
    ("/trust-rank", "What regulation binds you", "Shield, sovereignty, obligations register", "AIE, postures", 9),
    ("/decision-desk", "The call you must defend", "AIE data, cited corpus, live analyst", "AIE LIVE, AIE, SAMPLE", 5),
    ("/news-feed", "What moved", "/api/news, 24 h TTL", "AIE LIVE, AIE, LIVE", 0),
    ("/company-view (Your AI Position)", "Where we stand", "Live web research per company", "LIVE, AIE LIVE", 13),
    ("/admin", "What the catalogue holds and what it cost", "catalogue tables, cost model", "LIVE, DERIVED", 0),
]
r = 5
for p, q, d, lanes, sample in PAGES:
    write(ws4, r, [p, q, d, lanes, sample], fills={5: BAD if sample else OK}, wrap_cols=(2, 3, 4))
    ws4.cell(row=r, column=1).font = MONO
    ws4.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="top")
    ws4.row_dimensions[r].height = 30
    r += 1
ws4.cell(row=r + 1, column=1, value="Total SAMPLE badges across the product").font = BOLD
tot = ws4.cell(row=r + 1, column=5, value=f"=SUM(E5:E{r - 1})")
tot.font = BOLD
tot.alignment = Alignment(horizontal="center")
widths(ws4, {"A": 32, "B": 34, "C": 48, "D": 26, "E": 16})

# ─── 5. Refresh commands ────────────────────────────────────────────────────
ws5 = wb.create_sheet("Refreshing")
title(ws5, "How to refresh each source", "Cost is measured from lib/admin/cost-model.ts at list prices, not estimated.")
header(ws5, 4, ["What", "Command", "Writes to", "Needs", "Cost per run"])
CMDS = [
    ("SEC disclosure snapshot", "npm run ingest:adoption", "data/adoption/*.json (must be committed to take effect)", "SEC_USER_AGENT (has a default)", 0.000019),
    ("Catalogue, all series", "npm run ingest:catalogue", "Supabase Postgres, immediate", "SUPABASE_SERVICE_ROLE_KEY. Never in the app runtime", 0.000022),
    ("AIE fixtures", "node scripts/sync-aie-fixtures.mjs", "fixtures/aie-live/", "Nothing", 0.000007),
    ("Category assessment", "node scripts/sync-category-rankings.mjs", "fixtures/aie-live/category-rankings.json", "Nothing. Parses v1's published pages", 0.000007),
    ("Private-company figures", "manual, cited per row", "lib/finance/data/private-figures.json", "A primary source per figure", 0.000006),
    ("News cache", "automatic", "24 h module TTL", "Nothing", 0.000026),
]
r = 5
for what, cmd, writes, needs, cost in CMDS:
    write(ws5, r, [what, cmd, writes, needs, cost], wrap_cols=(3, 4))
    ws5.cell(row=r, column=1).font = BOLD
    ws5.cell(row=r, column=2).font = MONO
    c = ws5.cell(row=r, column=5)
    c.number_format = "$#,##0.000000"
    ws5.row_dimensions[r].height = 30
    r += 1
ws5.cell(row=r, column=1, value="All series, every day, for a month").font = BOLD
t = ws5.cell(row=r, column=5, value=f"=SUM(E5:E{r - 1})*30")
t.font = BOLD
t.number_format = "$#,##0.0000"
ws5.cell(row=r + 1, column=1, value="Cost is not a reason to refresh less often. Rate limits and source courtesy are: the SEC fair-access header is a real obligation.").font = SUB_FONT
ws5.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
widths(ws5, {"A": 28, "B": 40, "C": 44, "D": 40, "E": 16})

# ─── 6. Environment ─────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Environment")
title(ws6, "Environment variables", "Secrets stay server-side. The one publishable key is marked as such and is protected by row-level security, not by secrecy.")
header(ws6, 4, ["Variable", "Needed for", "Secret?", "Notes"])
ENV = [
    ("ANALYSTGENIUS_API_BASE", "BoardRadar", "No", "Upstream base URL"),
    ("ANALYSTGENIUS_API_KEY", "BoardRadar", "YES", "Injected server-side only. Never reaches the browser"),
    ("NEXT_PUBLIC_SUPABASE_URL", "Catalogue reads", "No", "Defaulted"),
    ("NEXT_PUBLIC_SUPABASE_ANON_KEY", "Catalogue reads", "No", "Publishable by design. Ships to browsers. RLS is the protection"),
    ("SUPABASE_SERVICE_ROLE_KEY", "ingest:catalogue only", "YES", "The only place it is used. Must never be in the app runtime"),
    ("ANTHROPIC_API_KEY", "Analyst and Interrogate", "YES", "Unset means scripted mode and zero LLM spend"),
    ("SEC_USER_AGENT", "SEC fair access", "No", "Has a working default. A real obligation, not a courtesy"),
    ("MOCK_MODE", "Demos without network", "No", "true serves fixtures everywhere"),
    ("DEMO_USER / DEMO_PASS", "Local basic auth", "YES", "Deliberately UNSET in Vercel so the demo is shareable by link. Do not 'fix' this"),
]
r = 5
for v, need, secret, note in ENV:
    write(ws6, r, [v, need, secret, note], fills={3: BAD if secret == "YES" else OK}, wrap_cols=(4,))
    ws6.cell(row=r, column=1).font = MONO
    ws6.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="top")
    ws6.row_dimensions[r].height = 28
    r += 1
widths(ws6, {"A": 34, "B": 26, "C": 12, "D": 62})

wb.save(OUT)
print(f"wrote {OUT}")
print(f"sheets: {', '.join(wb.sheetnames)}")
